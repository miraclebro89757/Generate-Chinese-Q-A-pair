#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import time
from openpyxl import load_workbook

def check_progress(filename: str = "chinese_qa_50000.xlsx"):
    """Check the progress of Q&A generation."""
    
    if not os.path.exists(filename):
        print(f"File '{filename}' does not exist yet.")
        return
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Count rows (excluding header)
        total_rows = ws.max_row - 1  # Subtract header row
        
        print(f"Progress Check for {filename}")
        print("=" * 50)
        print(f"Current Q&A pairs: {total_rows:,}")
        print(f"Target: 50,000")
        print(f"Progress: {total_rows/50000*100:.1f}%")
        print(f"Remaining: {50000 - total_rows:,}")
        
        if total_rows > 0:
            # Show some sample data
            print(f"\nSample Q&A pairs:")
            print("-" * 50)
            for row in range(2, min(7, ws.max_row + 1)):
                question = ws.cell(row=row, column=1).value
                answer_type = ws.cell(row=row, column=2).value
                answer = ws.cell(row=row, column=3).value
                
                print(f"{row-1:2d}. 问题: {question}")
                print(f"    回答类型: {answer_type}")
                print(f"    回答: {answer[:50]}{'...' if len(answer) > 50 else ''}")
                print()
        
        # File size info
        file_size = os.path.getsize(filename) / (1024 * 1024)  # MB
        print(f"File size: {file_size:.2f} MB")
        
    except Exception as e:
        print(f"Error reading file: {e}")

def main():
    """Main function to check progress."""
    print("Q&A Generation Progress Checker")
    print("=" * 50)
    
    filename = "chinese_qa_50000.xlsx"
    check_progress(filename)
    
    print(f"\nTo check progress again, run: python3 check_progress.py")

if __name__ == "__main__":
    main() 
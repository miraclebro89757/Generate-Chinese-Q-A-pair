#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import time

class FixedSizeExcelGenerator:
    def __init__(self):
        self.chinese_words = [
            "人工智能", "机器学习", "深度学习", "大数据", "云计算", "区块链", "物联网", "5G技术",
            "虚拟现实", "增强现实", "自动驾驶", "机器人", "无人机", "3D打印", "量子计算", "生物技术",
            "新能源", "环保技术", "智慧城市", "数字孪生", "边缘计算", "容器技术", "微服务", "API",
            "网络安全", "数据隐私", "密码学", "分布式系统", "高并发", "负载均衡", "缓存策略", "数据库优化"
        ]
        
        self.verbs = ["实现", "优化", "部署", "维护", "扩展", "测试", "监控", "升级"]
        self.adjectives = ["高效的", "可靠的", "安全的", "快速的", "智能的", "创新的", "先进的", "稳定的"]
        self.technologies = ["机器学习算法", "深度学习模型", "神经网络", "自然语言处理", "计算机视觉"]

    def generate_random_message(self) -> str:
        """Generate a random Chinese message."""
        patterns = [
            f"{random.choice(self.adjectives)}{random.choice(self.chinese_words)}",
            f"{random.choice(self.chinese_words)}的{random.choice(self.verbs)}",
            f"{random.choice(self.adjectives)}{random.choice(self.chinese_words)}通过{random.choice(self.technologies)}实现{random.choice(self.verbs)}",
            f"{random.choice(self.chinese_words)}利用{random.choice(self.technologies)}进行{random.choice(self.verbs)}"
        ]
        
        message = random.choice(patterns)
        if random.random() > 0.5:
            message += f"，这种{random.choice(self.adjectives)}{random.choice(self.chinese_words)}技术具有{random.choice(['高效性', '可靠性', '安全性'])}。"
        
        return message + "。"

    def create_excel_with_size(self, target_size_mb: float, filename: str = None):
        """Create an Excel file with random messages to reach target size."""
        
        if filename is None:
            filename = f"random_messages_{target_size_mb}MB.xlsx"
        
        print(f"Generating Excel file with target size: {target_size_mb}MB")
        print(f"Target file: {filename}")
        print("=" * 60)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "随机消息数据"
        
        # Write headers
        headers = ["消息ID", "消息内容", "消息类型", "时间戳", "优先级", "来源", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        current_row = 2
        total_messages = 0
        start_time = time.time()
        
        while True:
            # Generate batch of messages
            batch_size = 1000
            for _ in range(batch_size):
                message = self.generate_random_message()
                
                ws.cell(row=current_row, column=1, value=f"MSG_{random.randint(10000, 99999)}")
                ws.cell(row=current_row, column=2, value=message)
                ws.cell(row=current_row, column=3, value=random.choice(["信息", "警告", "错误", "成功", "提示"]))
                ws.cell(row=current_row, column=4, value=f"2024-{random.randint(1, 12):02d}-{random.randint(1, 28):02d} {random.randint(0, 23):02d}:{random.randint(0, 59):02d}:{random.randint(0, 59):02d}")
                ws.cell(row=current_row, column=5, value=random.choice(["高", "中", "低"]))
                ws.cell(row=current_row, column=6, value=random.choice(["系统", "用户", "应用", "服务", "数据库"]))
                ws.cell(row=current_row, column=7, value=random.choice(["活跃", "待处理", "已完成", "已取消", "暂停"]))
                current_row += 1
            
            total_messages += batch_size
            
            # Save and check file size
            wb.save(filename)
            file_size_mb = os.path.getsize(filename) / (1024 * 1024)
            
            # Progress update
            elapsed_time = time.time() - start_time
            print(f"Generated {total_messages:,} messages, File size: {file_size_mb:.2f}MB ({file_size_mb/target_size_mb*100:.1f}%)")
            
            # Check if target size reached
            if file_size_mb >= target_size_mb:
                break
        
        total_time = time.time() - start_time
        final_size_mb = os.path.getsize(filename) / (1024 * 1024)
        
        print(f"\n" + "=" * 60)
        print(f"GENERATION COMPLETED!")
        print(f"Total messages: {total_messages:,}")
        print(f"Final file size: {final_size_mb:.2f}MB")
        print(f"Target size: {target_size_mb:.2f}MB")
        print(f"Accuracy: {final_size_mb/target_size_mb*100:.1f}%")
        print(f"Total time: {total_time/60:.1f} minutes")
        print(f"File saved as: {filename}")
        
        return total_messages, final_size_mb

def main():
    """Main function for fixed-size Excel generation."""
    print("Fixed Size Excel Generator")
    print("=" * 60)
    
    # Configuration - Change target_size_mb to 5.0, 20.0, etc.
    target_size_mb = 5.0  # Change this to your desired size
    filename = f"random_messages_{target_size_mb}MB.xlsx"
    
    print(f"Configuration:")
    print(f"- Target size: {target_size_mb}MB")
    print(f"- Output file: {filename}")
    
    generator = FixedSizeExcelGenerator()
    total_messages, final_size = generator.create_excel_with_size(target_size_mb, filename)

if __name__ == "__main__":
    main()

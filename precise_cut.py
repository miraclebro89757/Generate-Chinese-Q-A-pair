#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os
import shutil

def precise_cut_excel_size(input_filename: str, target_size_mb: float = 20.0):
    """Precisely cut Excel file size to target size with frequent checking."""
    
    if not os.path.exists(input_filename):
        print(f"Error: File '{input_filename}' does not exist.")
        return
    
    # Create backup
    backup_filename = input_filename.replace('.xlsx', '_backup.xlsx')
    shutil.copy2(input_filename, backup_filename)
    print(f"Created backup: {backup_filename}")
    
    # Load workbook
    wb = load_workbook(input_filename)
    ws = wb.active
    
    # Get current file size
    current_size_mb = os.path.getsize(input_filename) / (1024 * 1024)
    print(f"Current file size: {current_size_mb:.2f}MB")
    print(f"Target size: {target_size_mb:.2f}MB")
    
    if current_size_mb <= target_size_mb:
        print("File is already smaller than target size. No changes needed.")
        return
    
    total_rows = ws.max_row
    header_row = 1
    data_rows = total_rows - header_row
    
    print(f"Total rows: {total_rows} (including header)")
    print(f"Data rows: {data_rows}")
    
    # Remove rows one by one and check size frequently
    removed_count = 0
    check_interval = 1000  # Check size every 1000 rows
    
    while ws.max_row > header_row + 1:
        # Remove one row from bottom
        ws.delete_rows(ws.max_row)
        removed_count += 1
        
        # Check size periodically
        if removed_count % check_interval == 0:
            wb.save(input_filename)
            current_size = os.path.getsize(input_filename) / (1024 * 1024)
            print(f"Removed {removed_count} rows, Current size: {current_size:.2f}MB")
            
            if current_size <= target_size_mb:
                print("Target size reached!")
                break
    
    # Final save
    wb.save(input_filename)
    
    # Check final size
    final_size_mb = os.path.getsize(input_filename) / (1024 * 1024)
    final_rows = ws.max_row
    
    print(f"\n" + "=" * 50)
    print(f"PRECISE FILE SIZE REDUCTION COMPLETED!")
    print(f"Original size: {current_size_mb:.2f}MB")
    print(f"Final size: {final_size_mb:.2f}MB")
    print(f"Size reduction: {current_size_mb - final_size_mb:.2f}MB")
    print(f"Original rows: {total_rows}")
    print(f"Final rows: {final_rows}")
    print(f"Rows removed: {total_rows - final_rows}")
    print(f"Target achieved: {'Yes' if final_size_mb <= target_size_mb else 'No'}")
    print(f"File: {input_filename}")

def main():
    """Main function for precise file size cutting."""
    print("Precise Excel File Size Cutter")
    print("=" * 50)
    
    # Configuration
    input_filename = "random_messages_20.0MB.xlsx"
    target_size_mb = 19.0
    
    print(f"Input file: {input_filename}")
    print(f"Target size: {target_size_mb}MB")
    
    precise_cut_excel_size(input_filename, target_size_mb)

if __name__ == "__main__":
    main() 
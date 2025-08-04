#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os
import shutil

def cut_excel_file_size(input_filename: str, target_size_mb: float = 19.0):
    """Cut Excel file size to target size by removing rows."""
    
    if not os.path.exists(input_filename):
        print(f"Error: File '{input_filename}' does not exist.")
        return
    
    # Create backup
    backup_filename = input_filename.replace('.xlsx', '_backup.xlsx')
    shutil.copy2(input_filename, backup_filename)
    print(f"Created backup: {backup_filename}")
    
    # Load workbook
    wb = load_workbook(input_filename)
    ws = wb.active
    
    # Get current file size
    current_size_mb = os.path.getsize(input_filename) / (1024 * 1024)
    print(f"Current file size: {current_size_mb:.2f}MB")
    print(f"Target size: {target_size_mb:.2f}MB")
    
    if current_size_mb <= target_size_mb:
        print("File is already smaller than target size. No changes needed.")
        return
    
    # Calculate how many rows to remove
    total_rows = ws.max_row
    header_row = 1
    data_rows = total_rows - header_row
    
    print(f"Total rows: {total_rows} (including header)")
    print(f"Data rows: {data_rows}")
    
    # Estimate rows to remove (rough calculation)
    size_ratio = target_size_mb / current_size_mb
    target_data_rows = int(data_rows * size_ratio)
    rows_to_remove = data_rows - target_data_rows
    
    print(f"Estimated rows to remove: {rows_to_remove}")
    
    # Remove rows from bottom (keep header)
    if rows_to_remove > 0:
        print(f"Removing {rows_to_remove} rows from bottom...")
        
        # Remove rows in batches to avoid memory issues
        batch_size = 1000
        removed_count = 0
        
        while removed_count < rows_to_remove and ws.max_row > header_row + 1:
            current_batch = min(batch_size, rows_to_remove - removed_count, ws.max_row - header_row - 1)
            
            # Remove rows from bottom
            for _ in range(current_batch):
                ws.delete_rows(ws.max_row)
                removed_count += 1
            
            # Save periodically to check size
            if removed_count % 5000 == 0:
                wb.save(input_filename)
                current_size = os.path.getsize(input_filename) / (1024 * 1024)
                print(f"Removed {removed_count} rows, Current size: {current_size:.2f}MB")
                
                if current_size <= target_size_mb:
                    print("Target size reached!")
                    break
        
        # Final save
        wb.save(input_filename)
    
    # Check final size
    final_size_mb = os.path.getsize(input_filename) / (1024 * 1024)
    final_rows = ws.max_row
    
    print(f"\n" + "=" * 50)
    print(f"FILE SIZE REDUCTION COMPLETED!")
    print(f"Original size: {current_size_mb:.2f}MB")
    print(f"Final size: {final_size_mb:.2f}MB")
    print(f"Size reduction: {current_size_mb - final_size_mb:.2f}MB")
    print(f"Original rows: {total_rows}")
    print(f"Final rows: {final_rows}")
    print(f"Rows removed: {total_rows - final_rows}")
    print(f"Target achieved: {'Yes' if final_size_mb <= target_size_mb else 'No'}")
    print(f"File: {input_filename}")

def main():
    """Main function to cut file size."""
    print("Excel File Size Cutter")
    print("=" * 50)
    
    # Configuration
    input_filename = "random_messages_18.0MB.xlsx"
    target_size_mb = 18.0
    
    print(f"Input file: {input_filename}")
    print(f"Target size: {target_size_mb}MB")
    
    cut_excel_file_size(input_filename, target_size_mb)

if __name__ == "__main__":
    main() 
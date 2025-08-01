#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Chinese Q&A Generator
Generates random Chinese questions and answers, then writes them to Excel.
"""

import random
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
import re
import os

class ChineseQAGenerator:
    def __init__(self):
        self.used_questions = set()
        
        # More diverse question templates
        self.question_templates = [
            # Basic questions
            "什么是{}？", "{}是什么？", "{}的特点是什么？", "{}的作用是什么？",
            "{}的定义是什么？", "{}的分类有哪些？", "{}的历史是什么？", "{}的原理是什么？",
            "{}的优势是什么？", "{}的缺点是什么？", "{}的发展趋势是什么？", "{}的应用场景有哪些？",
            
            # More specific questions
            "{}如何工作？", "{}的工作原理是什么？", "{}的核心技术是什么？", "{}的关键要素是什么？",
            "{}的实现方式有哪些？", "{}的架构设计是什么？", "{}的性能指标是什么？", "{}的优化方法是什么？",
            "{}的部署流程是什么？", "{}的维护策略是什么？", "{}的扩展性如何？", "{}的安全性如何？",
            
            # Comparative questions
            "{}与{}有什么区别？", "{}相比{}有什么优势？", "{}和{}哪个更好？",
            "{}与{}的异同点是什么？", "{}相对于{}有什么特点？",
            
            # Process questions
            "如何实现{}？", "如何优化{}？", "如何部署{}？", "如何维护{}？",
            "如何扩展{}？", "如何测试{}？", "如何监控{}？", "如何升级{}？",
            
            # Problem-solving questions
            "{}常见问题有哪些？", "{}的故障排除方法是什么？", "{}的性能瓶颈在哪里？",
            "{}的安全风险是什么？", "{}的兼容性问题是什么？", "{}的扩展限制是什么？",
            
            # Future-oriented questions
            "{}的未来发展方向是什么？", "{}的技术演进趋势是什么？", "{}的市场前景如何？",
            "{}的替代方案有哪些？", "{}的升级路径是什么？", "{}的创新点在哪里？"
        ]
        
        # Expanded topics with more specific categories
        self.topics = {
            "AI_ML": [
                "机器学习算法", "深度学习模型", "神经网络", "自然语言处理", "计算机视觉",
                "强化学习", "迁移学习", "联邦学习", "图神经网络", "Transformer模型",
                "卷积神经网络", "循环神经网络", "生成对抗网络", "自编码器", "支持向量机",
                "决策树", "随机森林", "梯度提升", "聚类算法", "降维技术"
            ],
            "BigData": [
                "大数据处理", "数据挖掘", "数据仓库", "数据湖", "流数据处理",
                "批处理系统", "实时分析", "数据可视化", "数据治理", "数据质量",
                "数据安全", "数据隐私", "数据备份", "数据恢复", "数据迁移"
            ],
            "Cloud": [
                "云计算平台", "容器技术", "微服务架构", "服务网格", "无服务器计算",
                "云原生应用", "混合云", "多云管理", "云安全", "云监控",
                "云存储", "云数据库", "云网络", "云负载均衡", "云弹性伸缩"
            ],
            "DevOps": [
                "持续集成", "持续部署", "DevOps工具链", "自动化测试", "配置管理",
                "容器编排", "服务发现", "日志管理", "监控告警", "性能优化",
                "故障恢复", "蓝绿部署", "金丝雀发布", "滚动更新", "回滚策略"
            ],
            "Security": [
                "网络安全", "数据加密", "身份认证", "访问控制", "漏洞扫描",
                "入侵检测", "防火墙", "VPN技术", "零信任架构", "安全审计",
                "威胁情报", "安全运营", "应急响应", "合规管理", "风险评估"
            ],
            "Database": [
                "关系型数据库", "NoSQL数据库", "分布式数据库", "数据库优化", "索引策略",
                "事务管理", "并发控制", "数据备份", "数据恢复", "数据库监控",
                "数据库安全", "数据库迁移", "分库分表", "读写分离", "缓存策略"
            ],
            "Mobile": [
                "移动应用开发", "跨平台开发", "原生开发", "混合开发", "移动UI设计",
                "移动性能优化", "移动安全", "推送通知", "移动支付", "移动广告",
                "移动分析", "移动测试", "应用商店", "版本管理", "热更新"
            ],
            "Web": [
                "前端框架", "后端开发", "API设计", "RESTful接口", "GraphQL",
                "Web安全", "性能优化", "SEO优化", "响应式设计", "渐进式应用",
                "单页应用", "服务端渲染", "静态站点生成", "CDN加速", "缓存策略"
            ]
        }
        
        # Flatten topics for easier access
        self.all_topics = []
        for category_topics in self.topics.values():
            self.all_topics.extend(category_topics)
        
        self.answer_types = ["纯文本", "富文本"]
        
        # More complex answer patterns
        self.answer_patterns = [
            # Basic patterns
            "{}是一种{}技术，主要用于{}。",
            "{}指的是{}，具有{}的特点。",
            "{}的核心是{}，通过{}实现功能。",
            "{}包括{}，其中最重要的是{}。",
            "{}的发展经历了{}，目前处于{}阶段。",
            
            # Technical patterns
            "{}通过{}算法实现{}功能，能够{}。",
            "{}基于{}架构设计，采用{}技术，支持{}。",
            "{}利用{}原理，结合{}方法，实现{}。",
            "{}采用{}模式，集成{}组件，提供{}服务。",
            "{}运用{}策略，优化{}性能，提升{}效率。",
            
            # Process patterns
            "{}的实现过程包括{}、{}和{}三个主要步骤。",
            "{}的部署流程涉及{}配置、{}测试和{}监控。",
            "{}的维护工作包括{}检查、{}更新和{}优化。",
            "{}的扩展方案通过{}架构、{}技术和{}策略实现。",
            
            # Comparison patterns
            "{}相比{}具有{}优势，但在{}方面存在{}限制。",
            "{}与{}的主要区别在于{}，前者{}，后者{}。",
            "{}和{}各有特点，{}适合{}场景，{}适合{}场景。",
            
            # Problem-solving patterns
            "{}常见问题包括{}、{}和{}，解决方案分别是{}、{}和{}。",
            "{}的性能瓶颈主要在{}，可以通过{}、{}和{}方法优化。",
            "{}的安全风险包括{}、{}和{}，需要采取{}、{}和{}措施。",
            
            # Future patterns
            "{}的发展趋势是{}，未来将向{}方向发展，预计{}。",
            "{}的技术演进包括{}、{}和{}，将带来{}影响。",
            "{}的市场前景广阔，主要应用在{}、{}和{}领域。"
        ]
        
        # Answer components for more variety
        self.answer_components = {
            "技术": ["先进技术", "创新技术", "前沿技术", "核心技术", "基础技术", "成熟技术", "新兴技术"],
            "功能": ["数据处理", "信息传输", "智能分析", "自动化控制", "实时监控", "预测分析", "决策支持"],
            "特点": ["高效性", "可靠性", "可扩展性", "安全性", "易用性", "灵活性", "稳定性"],
            "阶段": ["起步阶段", "发展阶段", "成熟阶段", "创新阶段", "转型阶段", "优化阶段"],
            "影响": ["提高效率", "降低成本", "改善体验", "促进创新", "推动发展", "增强竞争力"],
            "方式": ["算法优化", "硬件升级", "软件改进", "架构重构", "流程优化", "策略调整"],
            "架构": ["分布式架构", "微服务架构", "云原生架构", "事件驱动架构", "分层架构"],
            "算法": ["机器学习算法", "深度学习算法", "优化算法", "搜索算法", "排序算法"],
            "协议": ["HTTP协议", "TCP协议", "WebSocket协议", "MQTT协议", "REST协议"],
            "标准": ["行业标准", "技术标准", "安全标准", "性能标准", "质量标准"]
        }

    def generate_unique_question(self) -> str:
        attempts = 0
        while attempts < 200:  # Increased attempts for more variety
            template = random.choice(self.question_templates)
            
            # Choose topics based on template type
            if "与" in template or "相比" in template:
                # For comparative questions, need two different topics
                topic1 = random.choice(self.all_topics)
                topic2 = random.choice(self.all_topics)
                while topic2 == topic1:
                    topic2 = random.choice(self.all_topics)
                try:
                    question = template.format(topic1, topic2)
                except IndexError:
                    # Fallback for templates that don't match expected format
                    question = template.format(topic1)
            else:
                topic = random.choice(self.all_topics)
                try:
                    question = template.format(topic)
                except IndexError:
                    # Fallback for templates that don't match expected format
                    question = f"什么是{topic}？"
            
            if question not in self.used_questions:
                self.used_questions.add(question)
                return question
            attempts += 1
        
        # Add random number and timestamp to make unique
        question = f"{question}（{random.randint(1, 99999)}-{random.randint(1000, 9999)}）"
        self.used_questions.add(question)
        return question

    def generate_answer(self, question: str) -> str:
        # Extract topic from question
        topic_match = re.search(r'[什么是如何与相比]*([^？\s]+)[？\s]', question)
        topic = topic_match.group(1) if topic_match else random.choice(self.all_topics)
        
        pattern = random.choice(self.answer_patterns)
        
        # Generate components based on pattern complexity
        components_needed = pattern.count('{}') - 1
        components = []
        
        for i in range(components_needed):
            if "技术" in pattern:
                components.append(random.choice(self.answer_components["技术"]))
            elif "功能" in pattern:
                components.append(random.choice(self.answer_components["功能"]))
            elif "特点" in pattern:
                components.append(random.choice(self.answer_components["特点"]))
            elif "阶段" in pattern:
                components.append(random.choice(self.answer_components["阶段"]))
            elif "影响" in pattern:
                components.append(random.choice(self.answer_components["影响"]))
            elif "方式" in pattern:
                components.append(random.choice(self.answer_components["方式"]))
            elif "架构" in pattern:
                components.append(random.choice(self.answer_components["架构"]))
            elif "算法" in pattern:
                components.append(random.choice(self.answer_components["算法"]))
            elif "协议" in pattern:
                components.append(random.choice(self.answer_components["协议"]))
            elif "标准" in pattern:
                components.append(random.choice(self.answer_components["标准"]))
            else:
                # Use a random topic from a different category
                category = random.choice(list(self.topics.keys()))
                components.append(random.choice(self.topics[category]))
        
        answer = pattern.format(topic, *components)
        
        # Ensure answer is within 200 characters
        if len(answer) > 200:
            answer = answer[:197] + "..."
        
        return answer

    def generate_qa_pairs(self, count: int) -> List[Dict[str, str]]:
        qa_pairs = []
        for i in range(count):
            question = self.generate_unique_question()
            answer = self.generate_answer(question)
            answer_type = random.choice(self.answer_types)
            
            qa_pairs.append({
                "标准问题": question,
                "回答类型": answer_type,
                "问题回答1": answer
            })
        return qa_pairs

    def load_existing_questions(self, filename: str) -> set:
        """Load existing questions from Excel file to avoid duplicates."""
        existing_questions = set()
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                ws = wb.active
                
                # Read existing questions from column A (starting from row 2)
                for row in range(2, ws.max_row + 1):
                    question = ws.cell(row=row, column=1).value
                    if question:
                        existing_questions.add(question)
                
                print(f"Loaded {len(existing_questions)} existing questions from {filename}")
            except Exception as e:
                print(f"Warning: Could not load existing questions from {filename}: {e}")
        
        return existing_questions

    def write_to_excel(self, qa_pairs: List[Dict[str, str]], filename: str = "chinese_qa_data.xlsx", append: bool = False):
        """Write Q&A pairs to Excel file with proper formatting."""
        
        if append and os.path.exists(filename):
            # Load existing workbook
            wb = load_workbook(filename)
            ws = wb.active
            
            # Load existing questions to avoid duplicates
            existing_questions = self.load_existing_questions(filename)
            self.used_questions.update(existing_questions)
            
            # Filter out questions that already exist
            new_qa_pairs = []
            for qa in qa_pairs:
                if qa["标准问题"] not in existing_questions:
                    new_qa_pairs.append(qa)
                else:
                    print(f"Skipping duplicate question: {qa['标准问题']}")
            
            if not new_qa_pairs:
                print("No new questions to add - all questions already exist in the file.")
                return
            
            qa_pairs = new_qa_pairs
            start_row = ws.max_row + 1
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "中文问答数据"
            
            # Write headers with formatting
            headers = ["标准问题 (必填)", "回答类型 (必填)", "问题回答1 (必填)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            start_row = 2
        
        # Write data
        for row_idx, qa_pair in enumerate(qa_pairs, start_row):
            ws.cell(row=row_idx, column=1, value=qa_pair["标准问题"])
            ws.cell(row=row_idx, column=2, value=qa_pair["回答类型"])
            ws.cell(row=row_idx, column=3, value=qa_pair["问题回答1"])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        action = "appended to" if append else "created"
        print(f"Excel file '{filename}' has been {action} successfully!")
        print(f"Added {len(qa_pairs)} new Q&A pairs.")

    def generate_and_save(self, count: int = 50, filename: str = "chinese_qa_data.xlsx", append: bool = False):
        print(f"Generating {count} unique Chinese Q&A pairs...")
        qa_pairs = self.generate_qa_pairs(count)
        
        print(f"Writing data to Excel file '{filename}'...")
        self.write_to_excel(qa_pairs, filename, append)
        
        print(f"Successfully generated {len(qa_pairs)} Q&A pairs!")
        return qa_pairs

def main():
    generator = ChineseQAGenerator()
    
    # Check if file exists to determine append mode
    filename = "chinese_qa_data.xlsx"
    append_mode = os.path.exists(filename)
    
    if append_mode:
        print(f"Found existing file '{filename}'. Will append new Q&A pairs.")
        qa_pairs = generator.generate_and_save(count=50, filename=filename, append=True)
    else:
        print(f"Creating new file '{filename}'.")
        qa_pairs = generator.generate_and_save(count=50, filename=filename, append=False)
    
    print("\nFirst 5 Q&A pairs:")
    for i, qa in enumerate(qa_pairs[:5], 1):
        print(f"\n{i}. 问题: {qa['标准问题']}")
        print(f"   回答类型: {qa['回答类型']}")
        print(f"   回答: {qa['问题回答1']}")

if __name__ == "__main__":
    main() 